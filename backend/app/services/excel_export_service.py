"""Servicio de exportación de operaciones financieras a Excel (.xlsx)."""

from datetime import date, datetime, timezone
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy.orm import Session

from app.models.operacion import Operacion, TipoOperacion
from app.models.distribucion import DistribucionDetalle
from app.models.area import Area
from app.models.socio import Socio

# ---------------------------------------------------------------------------
# Colores institucionales
# ---------------------------------------------------------------------------
AZUL_INSTITUCIONAL = "1E3A8A"
AZUL_HOVER = "1D4ED8"
AZUL_SOFT = "DBEAFE"
GRIS_BORDE = "E2E8F0"
GRIS_FONDO_ALT = "F8FAFC"
BLANCO = "FFFFFF"

# ---------------------------------------------------------------------------
# Estilos Big 4
# ---------------------------------------------------------------------------
header_font = Font(name="Calibri", size=10, bold=True, color=BLANCO)
header_fill = PatternFill(start_color=AZUL_INSTITUCIONAL, end_color=AZUL_INSTITUCIONAL, fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

data_font = Font(name="Calibri", size=10, color="334155")
data_font_usd = Font(name="Calibri", size=10, color=AZUL_INSTITUCIONAL, bold=True)
data_font_tc = Font(name="Calibri", size=10, color="64748B")

FORMAT_UYU = '#,##0.00'
FORMAT_USD = '#,##0.00'
FORMAT_TC = '#,##0.0000'
FORMAT_PCT = '0.00%'
FORMAT_DATE = 'DD/MM/YYYY'

thin_border = Border(bottom=Side(style="hair", color=GRIS_BORDE))
alt_fill = PatternFill(start_color=GRIS_FONDO_ALT, end_color=GRIS_FONDO_ALT, fill_type="solid")

# ---------------------------------------------------------------------------
# Mapeo de nombres canónicos
# ---------------------------------------------------------------------------
NOMBRE_CANONICO = {
    "Rodrigo Mederos": "R. Mederos",
    "Office 2000": "Office",
    "Club Cannabis": "Club Cannabico Estarse",
    "Mercurius Cesión Creditos Paigo": "Mercurius Fideicomiso",
    "Hipoteca J. Buriani": "Joaquin Buriani",
    "Camps": "Cams",
}


def _limpiar_texto(texto: str | None) -> str:
    """Retorna texto limpio (strip) o cadena vacía si None."""
    if texto is None:
        return ""
    return texto.strip()


def _normalizar_nombre(nombre: str | None) -> str:
    """strip + title + mapeo canónico. Retorna '' si None."""
    if nombre is None:
        return ""
    limpio = nombre.strip().title()
    return NOMBRE_CANONICO.get(limpio, limpio)


def _capitalizar_localidad(localidad) -> str:
    """Retorna localidad capitalizada (Montevideo / Mercedes)."""
    if localidad is None:
        return ""
    return localidad.value if hasattr(localidad, "value") else str(localidad).strip().title()


def _to_float(val) -> float | None:
    """Convierte Decimal/None a float para openpyxl."""
    if val is None:
        return None
    return float(val)


# ---------------------------------------------------------------------------
# Helpers de formato
# ---------------------------------------------------------------------------

# Mapping de columnas a tipo de formato
_COL_FORMATS = {
    "Fecha": "date",
    "T/C": "tc",
    "Monto Original": "money",
    "Monto UYU": "money",
    "Monto USD": "usd",
    "Total Pesificado": "money",
    "Total Dolarizado": "usd",
    "%": "pct",
}


def _apply_cell_style(cell, col_name: str, row_idx: int):
    """Aplica fuente, formato numérico y alineación según columna."""
    fmt = _COL_FORMATS.get(col_name)

    if fmt == "date":
        cell.number_format = FORMAT_DATE
        cell.font = data_font
    elif fmt == "tc":
        cell.number_format = FORMAT_TC
        cell.font = data_font_tc
        cell.alignment = Alignment(horizontal="right")
    elif fmt == "money":
        cell.number_format = FORMAT_UYU
        cell.font = data_font
        cell.alignment = Alignment(horizontal="right")
    elif fmt == "usd":
        cell.number_format = FORMAT_USD
        cell.font = data_font_usd
        cell.alignment = Alignment(horizontal="right")
    elif fmt == "pct":
        cell.number_format = FORMAT_PCT
        cell.font = data_font
        cell.alignment = Alignment(horizontal="right")
    else:
        cell.font = data_font

    cell.border = thin_border

    # Fila alternada (row_idx es 1-based en la hoja, datos arrancan en fila 2)
    if row_idx % 2 == 0:
        cell.fill = alt_fill


def _write_sheet(ws, headers: list[str], rows: list[list], col_widths_min: int = 12):
    """Escribe headers + datos con formato Big 4."""
    # Headers
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Datos
    for r_idx, row_data in enumerate(rows, 2):
        for c_idx, (val, col_name) in enumerate(zip(row_data, headers), 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            _apply_cell_style(cell, col_name, r_idx)

    # Auto-filter
    if rows:
        last_col_letter = chr(ord("A") + len(headers) - 1) if len(headers) <= 26 else "Z"
        ws.auto_filter.ref = f"A1:{last_col_letter}{len(rows) + 1}"
    else:
        last_col_letter = chr(ord("A") + len(headers) - 1) if len(headers) <= 26 else "Z"
        ws.auto_filter.ref = f"A1:{last_col_letter}1"

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-width
    for col_idx, h in enumerate(headers, 1):
        max_len = len(h)
        for row_data in rows:
            val = row_data[col_idx - 1]
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(max_len + 2, col_widths_min)


def _write_info_sheet(ws, info_rows: list[tuple[str, str]]):
    """Escribe la hoja Info Export en formato clave-valor."""
    label_font = Font(name="Calibri", size=10, bold=True, color=AZUL_INSTITUCIONAL)
    value_font = Font(name="Calibri", size=10, color="334155")
    for r_idx, (label, value) in enumerate(info_rows, 1):
        cell_a = ws.cell(row=r_idx, column=1, value=label)
        cell_a.font = label_font
        cell_b = ws.cell(row=r_idx, column=2, value=value)
        cell_b.font = value_font
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 50


# ---------------------------------------------------------------------------
# Función principal
# ---------------------------------------------------------------------------

def generar_excel_operaciones(
    db: Session,
    fecha_desde: date,
    fecha_hasta: date,
    usuario_nombre: str,
) -> bytes:
    """Genera un archivo Excel (.xlsx) con las operaciones financieras del período."""

    # 1. Consulta BD — operaciones activas en rango, con LEFT JOIN a areas
    operaciones = (
        db.query(Operacion, Area.nombre.label("area_nombre"))
        .outerjoin(Area, Operacion.area_id == Area.id)
        .filter(Operacion.deleted_at.is_(None))
        .filter(Operacion.fecha >= fecha_desde)
        .filter(Operacion.fecha <= fecha_hasta)
        .order_by(Operacion.fecha.desc())
        .all()
    )

    # 2. Separar por tipo
    ingresos = []
    gastos = []
    retiros = []
    distribuciones = []

    for op, area_nombre in operaciones:
        entry = (op, area_nombre)
        if op.tipo_operacion == TipoOperacion.INGRESO:
            ingresos.append(entry)
        elif op.tipo_operacion == TipoOperacion.GASTO:
            gastos.append(entry)
        elif op.tipo_operacion == TipoOperacion.RETIRO:
            retiros.append(entry)
        elif op.tipo_operacion == TipoOperacion.DISTRIBUCION:
            distribuciones.append(entry)

    # 3. Distribuciones detalle con JOIN a socios
    dist_ids = [op.id for op, _ in distribuciones]
    detalle_socios = []
    if dist_ids:
        detalle_socios = (
            db.query(
                DistribucionDetalle,
                Operacion.fecha,
                Operacion.localidad,
                Operacion.tipo_cambio,
                Socio.nombre.label("socio_nombre"),
            )
            .join(Operacion, DistribucionDetalle.operacion_id == Operacion.id)
            .join(Socio, DistribucionDetalle.socio_id == Socio.id)
            .filter(DistribucionDetalle.operacion_id.in_(dist_ids))
            .order_by(Operacion.fecha.desc(), Socio.nombre)
            .all()
        )

    # 4. Preparar filas para cada hoja

    # Hoja Ingresos
    ingresos_headers = [
        "Fecha", "Área", "Localidad", "Cliente", "Descripción",
        "Moneda", "Monto Original", "T/C", "Monto UYU", "Monto USD",
        "Total Pesificado", "Total Dolarizado",
    ]
    ingresos_rows = []
    for op, area_nombre in ingresos:
        ingresos_rows.append([
            op.fecha,
            area_nombre or "",
            _capitalizar_localidad(op.localidad),
            _normalizar_nombre(op.cliente),
            _limpiar_texto(op.descripcion),
            op.moneda_original.value if op.moneda_original else "",
            _to_float(op.monto_original),
            _to_float(op.tipo_cambio),
            _to_float(op.monto_uyu),
            _to_float(op.monto_usd),
            _to_float(op.total_pesificado),
            _to_float(op.total_dolarizado),
        ])

    # Hoja Gastos
    gastos_headers = [
        "Fecha", "Área", "Localidad", "Proveedor", "Descripción",
        "Moneda", "Monto Original", "T/C", "Monto UYU", "Monto USD",
        "Total Pesificado", "Total Dolarizado",
    ]
    gastos_rows = []
    for op, area_nombre in gastos:
        gastos_rows.append([
            op.fecha,
            area_nombre or "",
            _capitalizar_localidad(op.localidad),
            _normalizar_nombre(op.proveedor),
            _limpiar_texto(op.descripcion),
            op.moneda_original.value if op.moneda_original else "",
            _to_float(op.monto_original),
            _to_float(op.tipo_cambio),
            _to_float(op.monto_uyu),
            _to_float(op.monto_usd),
            _to_float(op.total_pesificado),
            _to_float(op.total_dolarizado),
        ])

    # Hoja Retiros
    retiros_headers = [
        "Fecha", "Localidad", "Descripción", "T/C",
        "Monto UYU", "Monto USD", "Total Pesificado", "Total Dolarizado",
    ]
    retiros_rows = []
    for op, _ in retiros:
        retiros_rows.append([
            op.fecha,
            _capitalizar_localidad(op.localidad),
            _limpiar_texto(op.descripcion),
            _to_float(op.tipo_cambio),
            _to_float(op.monto_uyu),
            _to_float(op.monto_usd),
            _to_float(op.total_pesificado),
            _to_float(op.total_dolarizado),
        ])

    # Hoja Distribuciones
    distribuciones_headers = [
        "Fecha", "Localidad", "Descripción", "T/C",
        "Monto UYU", "Monto USD", "Total Pesificado", "Total Dolarizado",
    ]
    distribuciones_rows = []
    for op, _ in distribuciones:
        distribuciones_rows.append([
            op.fecha,
            _capitalizar_localidad(op.localidad),
            _limpiar_texto(op.descripcion),
            _to_float(op.tipo_cambio),
            _to_float(op.monto_uyu),
            _to_float(op.monto_usd),
            _to_float(op.total_pesificado),
            _to_float(op.total_dolarizado),
        ])

    # Hoja Detalle por Socio
    detalle_headers = [
        "Fecha", "Localidad", "Socio", "T/C",
        "Monto UYU", "Monto USD", "%", "Total Pesificado", "Total Dolarizado",
    ]
    detalle_rows = []
    for det, fecha, localidad, tipo_cambio, socio_nombre in detalle_socios:
        detalle_rows.append([
            fecha,
            _capitalizar_localidad(localidad),
            socio_nombre or "",
            _to_float(tipo_cambio),
            _to_float(det.monto_uyu),
            _to_float(det.monto_usd),
            _to_float(det.porcentaje) / 100.0 if det.porcentaje is not None else None,
            _to_float(det.total_pesificado),
            _to_float(det.total_dolarizado),
        ])

    # 5. Generar workbook
    wb = Workbook()

    # Hoja 1: Ingresos (renombrar la default)
    ws_ingresos = wb.active
    ws_ingresos.title = "Ingresos"
    _write_sheet(ws_ingresos, ingresos_headers, ingresos_rows)

    # Hoja 2: Gastos
    ws_gastos = wb.create_sheet("Gastos")
    _write_sheet(ws_gastos, gastos_headers, gastos_rows)

    # Hoja 3: Retiros
    ws_retiros = wb.create_sheet("Retiros")
    _write_sheet(ws_retiros, retiros_headers, retiros_rows)

    # Hoja 4: Distribuciones
    ws_dist = wb.create_sheet("Distribuciones")
    _write_sheet(ws_dist, distribuciones_headers, distribuciones_rows)

    # Hoja 5: Detalle por Socio
    ws_detalle = wb.create_sheet("Detalle por Socio")
    _write_sheet(ws_detalle, detalle_headers, detalle_rows)

    # Hoja 6: Info Export
    ws_info = wb.create_sheet("Info Export")
    total_ops = len(ingresos) + len(gastos) + len(retiros) + len(distribuciones)
    _write_info_sheet(ws_info, [
        ("Sistema", "CFO Inteligente — Conexion Consultora"),
        ("Fecha de exportacion", datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M UTC")),
        ("Exportado por", usuario_nombre),
        ("Periodo", f"{fecha_desde.strftime('%d/%m/%Y')} — {fecha_hasta.strftime('%d/%m/%Y')}"),
        ("Ingresos", f"{len(ingresos)} registros"),
        ("Gastos", f"{len(gastos)} registros"),
        ("Retiros", f"{len(retiros)} registros"),
        ("Distribuciones", f"{len(distribuciones)} registros"),
        ("Detalle por Socio", f"{len(detalle_rows)} registros"),
        ("Total operaciones", f"{total_ops} registros activos"),
    ])

    # 6. Guardar en memoria
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
